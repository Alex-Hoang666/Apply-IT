import tkinter as tk
import pandas as pd
# Thư viện để tạo ra một cửa sổ thông báo
from tkinter import END, messagebox
# thư viện để kiểm tra Mật Khẩu
import re
# Thư viện để Lưu thông tin tài khoản
from openpyxl import load_workbook


# Class của cửa sổ đăng nhập
## Sẽ thêm vào đăng nhập của quản lý hay nhân viên bán hàng ##
class LoginWindow:
    def __init__(self):
        self.window = tk.Tk()
         #lấy vị trí màn hình
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()

        # Tính toán vị trí giữa màn hình và hiện cửa sổ 1 ở chính giữa
        x = int((screen_width - 500) / 2)
        y = int((screen_height - 500) / 2)
        self.window.geometry(f"500x300+{x}+{y}")
        self.window.title("Đăng ký và Đăng nhập")

        #hiển thị mật khẩu
        self.show_password = tk.BooleanVar()
        self.show_password.set(False)

        self.label_username = tk.Label(self.window, text="Tên đăng nhập:")
        self.label_username.pack()
        self.entry_username = tk.Entry(self.window)
        self.entry_username.pack()

        self.label_password = tk.Label(self.window, text="Mật khẩu:")
        self.label_password.pack()
        self.entry_password = tk.Entry(self.window, show="*")
        self.entry_password.pack()

        self.button_register = tk.Button(self.window, text="Đăng ký", command=self.open_register_window)
        self.button_register.pack()

        self.button_login = tk.Button(self.window, text="Đăng nhập", command=self.login)
        self.button_login.pack()

        self.checkbutton_show_password = tk.Checkbutton(self.window, text="Hiển thị mật khẩu",variable=self.show_password,command=self.toggle_password_visibility)
        self.checkbutton_show_password.pack()

    # Hàm xử lý khi nút đăng kí được nhấn
    def open_register_window(self):
        register_window = RegisterWindow(self.window)

    # Hàm xử lý trình đăng nhập
    def login(self):
        username = self.entry_username.get()
        password = self.entry_password.get()
        # Xử lý logic đăng nhập tài khoản ở đây
        ##Sẽ thêm vào xử lý khi đăng nhập là quản lý và đăng nhập là nhân viên bán hàng dưới 2 luồn class khác nhau ##
        messagebox.showinfo("Thông báo", "Đăng nhập thành công!")

    # Hàm để khởi tạo cửa số GUI
    def run(self):
        self.window.mainloop()

    # Hàm hiện mật khẩu
    def toggle_password_visibility(self):
        if self.show_password.get():
            self.entry_password.config(show="")
        else:
            self.entry_password.config(show="*")

# Class của cửa sổ đăng kí
class RegisterWindow:
    def __init__(self, parent):
        self.register_window = tk.Toplevel(parent)
         #lấy vị trí màn hình
        screen_width = self.register_window.winfo_screenwidth()
        screen_height = self.register_window.winfo_screenheight()

        # Tính toán vị trí giữa màn hình và hiện cửa sổ 1 ở chính giữa
        x = int((screen_width - 500) / 2)
        y = int((screen_height - 500) / 2)
        self.register_window.geometry(f"300x200+{x}+{y}")
        self.register_window.title("Đăng ký tài khoản")

        # Hiển thị mật khẩu
        self.show_password = tk.BooleanVar()
        self.show_password.set(False)

        self.label_username = tk.Label(self.register_window, text="Tên đăng nhập:")
        self.label_username.pack()
        self.entry_username = tk.Entry(self.register_window)
        self.entry_username.pack()

        self.label_password = tk.Label(self.register_window, text="Mật khẩu:")
        self.label_password.pack()
        self.entry_password = tk.Entry(self.register_window, show="*")
        self.entry_password.pack()

        self.label_confirm_password = tk.Label(self.register_window, text="Nhập lại mật khẩu:")
        self.label_confirm_password.pack()
        self.entry_confirm_password = tk.Entry(self.register_window, show="*")
        self.entry_confirm_password.pack()

        self.label_error = tk.Label(self.register_window, text="")
        self.label_error.pack()

        self.button_register = tk.Button(self.register_window, text="Đăng ký", command=self.register)
        self.button_register.pack()

        self.checkbutton_show_password = tk.Checkbutton(self.register_window, text="Hiển thị mật khẩu",variable=self.show_password,command=self.toggle_password_visibility)
        self.checkbutton_show_password.pack()
    # Hàm hiện mật khẩu
    def toggle_password_visibility(self):
        if self.show_password.get():
            self.entry_password.config(show="")
            self.entry_confirm_password.config(show="")
        else:
            self.entry_password.config(show="*")
            self.entry_confirm_password.config(show="*")
    # hàm xử lý trình đăng kí 
    def register(self):
        username = self.entry_username.get()
        password = self.entry_password.get()
        confirm_password = self.entry_confirm_password.get()

        #kiểm tra xem người dùng nhập đủ thông tin chưa
        if not username or not password or not confirm_password:
            self.label_error.configure(text='ERROR: Vui lòng nhập đầy đủ thông tin!', fg="red")
            return

        # Kiểm tra nhập lại mật khẩu khớp không
        if password != confirm_password:
            self.label_error.configure(text='ERROR: Mật khẩu không khớp!', fg="red")
            self.entry_username.delete(0, END)
            self.entry_password.delete(0, END)
            self.entry_confirm_password.delete(0, END)
        else:
            #Kiểm tra mật khẩu hợp lệ
            if not self.validate_password(password):
                self.label_error.configure(text='ERROR: Mật khẩu không hợp lệ!', fg="red")
                self.entry_username.delete(0, END)
                self.entry_password.delete(0, END)
                self.entry_confirm_password.delete(0, END)
            # Kiểm tra tên đăng nhập tồn tại
            if self.check_username_exists(username):
                self.label_error.configure(text='ERROR: Tên đăng nhập đã tồn tại!', fg="red")
                self.entry_username.delete(0, END)
                self.entry_password.delete(0, END)
                self.entry_confirm_password.delete(0, END)
            else:
                # Xử lý logic đăng ký tài khoản ở đây
                self.update_excel(username, password)
                messagebox.showinfo("Thông báo", "Đăng ký thành công!")

    # Hàm kiểm tra các yêu cầu về mật khẩu (độ dài tối thiểu, chữ hoa, chữ thường, chữ số, ký tự đặc biệt)
    def validate_password(self, password):
        # Sử dụng biểu thức chính quy (regular expression)
        if len(password) < 8:
            return False
        if not re.search(r"[a-z]", password):
            return False
        if not re.search(r"\d", password):
            return False
        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
            return False
        return True
    
    #Hàm kiểm tra tên đăng nhập
    def check_username_exists(self, username):
        # Đường dẫn đến file Excel
        excel_file = "Du_Lieu_Dang_Nhap.xlsx"
        try:
            # Đọc dữ liệu từ file Excel
            data_frame = pd.read_excel(excel_file)

            # Kiểm tra xem tên đăng nhập đã tồn tại trong cột "Username" hay chưa
            if username in data_frame["Username"].values:
                return True

        except Exception as e:
            print("Lỗi khi đọc file Excel:", str(e))
        return False
    
    #Hàm lưu thông tin đăng nhập vào file excel
    def update_excel(self, username, password):
        # Mở tệp Excel hiện có
        workbook = load_workbook('Du_Lieu_Dang_Nhap.xlsx')
        
        # Chọn trang tính đầu tiên
        sheet = workbook.active
        
        # Tìm dòng cuối cùng có dữ liệu
        last_row = sheet.max_row + 1
        
        # Ghi thông tin đăng nhập vào tệp Excel
        sheet.cell(row=last_row, column=1).value = username
        sheet.cell(row=last_row, column=2).value = password
        
        # Lưu tệp Excel
        workbook.save('Du_Lieu_Dang_Nhap.xlsx')

            
#%%
# Tạo đối tượng cửa sổ đăng nhập và chạy ứng dụng
login_window = LoginWindow()
login_window.run()